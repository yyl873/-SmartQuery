"""
SmartQuery API 主入口
提供 NL2SQL 查询、数据库连接管理、CSV 导入/导出、表结构预览
"""

import os
import csv
import io
import json
import logging
import time
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse
from app.models import (
    QueryRequest, QueryResponse,
    GenerateRequest, GenerateResponse,
    Translation, ConnectRequest, TableInfo,
)
from app.database import (
    get_schema, execute_sql, import_csv, drop_table,
    reconnect, test_connection, get_connection_info,
)
from app.llm import generate_sql, generate_translation, fix_sql
from app.utils import is_safe_sql, format_sql, detect_encoding, is_safe_table_name

# ---- 日志配置 ----
LOG_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "logs")
os.makedirs(LOG_DIR, exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.StreamHandler(),                                        # 控制台输出
        logging.FileHandler(                                            # 文件输出
            os.path.join(LOG_DIR, "smartquery.log"),
            encoding="utf-8",
        ),
    ],
)
logger = logging.getLogger("smartquery")

# 创建 FastAPI 应用
app = FastAPI(
    title="SmartQuery API",
    description="NL2SQL 智能数据库查询助手 — 连接任意数据库，自然语言生成 SQL",
)

# CORS 中间件
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 静态文件
static_dir = os.path.join(os.path.dirname(__file__), "static")
app.mount("/static", StaticFiles(directory=static_dir), name="static")


# ==================== 页面 ====================

@app.get("/")
async def root():
    """返回 Web 界面"""
    return FileResponse(os.path.join(static_dir, "index.html"))


# ==================== 数据库连接管理 ====================

@app.post("/test-connection")
async def test_db_connection(req: ConnectRequest):
    """
    测试数据库连接是否可用（不切换当前连接）。
    返回 { "ok": true/false, "message": "..." }
    """
    logger.info(f"测试连接: {req.db_type}://{req.host or 'sqlite'}/{req.database}")
    try:
        ok = test_connection(
            db_type=req.db_type,
            host=req.host,
            port=req.port,
            database=req.database,
            user=req.user,
            password=req.password,
        )
        if ok:
            logger.info("连接测试成功")
            return {"ok": True, "message": "✅ 连接测试成功"}
        else:
            logger.warning("连接测试失败")
            return {"ok": False, "message": "❌ 连接失败，请检查参数"}
    except Exception as e:
        logger.error(f"连接测试异常: {e}")
        return {"ok": False, "message": f"❌ 连接异常: {str(e)}"}


@app.post("/connect")
async def connect_database(req: ConnectRequest):
    """
    切换到新的数据库连接。
    成功后返回当前所有表名。
    """
    logger.info(f"切换连接: {req.db_type}://{req.host or 'sqlite'}/{req.database}")
    try:
        reconnect(
            db_type=req.db_type,
            host=req.host,
            port=req.port,
            database=req.database,
            user=req.user,
            password=req.password,
        )
        info = get_connection_info()
        from app.database import engine
        from sqlalchemy import inspect
        inspector = inspect(engine)
        tables = inspector.get_table_names()
        logger.info(f"连接成功，{len(tables)} 张表")
        return {
            "ok": True,
            "message": f"✅ 已连接到 {req.db_type} 数据库",
            "connection": info,
            "tables": tables,
        }
    except Exception as e:
        logger.error(f"连接失败: {e}")
        raise HTTPException(status_code=400, detail=f"连接失败: {str(e)}")


@app.post("/disconnect")
async def disconnect_database():
    """
    断开当前数据库连接，回到默认 SQLite。
    """
    logger.info("断开连接，恢复到默认 SQLite")
    try:
        from app.config import DATABASE_URL
        reconnect(db_type="sqlite", database="smartquery.db")
        info = get_connection_info()
        from app.database import engine
        from sqlalchemy import inspect
        inspector = inspect(engine)
        tables = inspector.get_table_names()
        return {
            "ok": True,
            "message": "✅ 已断开，恢复到默认 SQLite 数据库",
            "connection": info,
            "tables": tables,
        }
    except Exception as e:
        logger.error(f"断开失败: {e}")
        raise HTTPException(status_code=500, detail=f"断开失败: {str(e)}")


@app.get("/connection-info")
async def current_connection():
    """获取当前数据库连接信息"""
    from app.database import engine
    from sqlalchemy import inspect
    info = get_connection_info()
    inspector = inspect(engine)
    tables = inspector.get_table_names()
    return {
        "connection": info,
        "tables": tables,
    }


# ==================== 表结构预览 ====================

@app.get("/table-info/{table_name}")
async def table_info(table_name: str):
    """
    返回指定表的结构信息（列名、类型、是否可空、默认值）。
    用于前端点击表名时预览。
    """
    from app.database import engine
    from sqlalchemy import inspect
    try:
        inspector = inspect(engine)
        if table_name not in inspector.get_table_names():
            raise HTTPException(status_code=404, detail=f"表 '{table_name}' 不存在")

        columns = inspector.get_columns(table_name)
        col_info = []
        for col in columns:
            col_info.append({
                "name": col["name"],
                "type": str(col["type"]),
                "nullable": col.get("nullable", True),
                "default": str(col.get("default")) if col.get("default") is not None else None,
                "primary_key": col.get("primary_key", False),
            })
        return {
            "table_name": table_name,
            "columns": col_info,
            "row_count": len(col_info),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取表结构失败: {str(e)}")


# ==================== SQL 生成（不执行） ====================

@app.post("/generate-sql", response_model=GenerateResponse)
async def generate_sql_only(request: GenerateRequest):
    """
    仅生成 SQL 语句，不执行。
    用户可在前端编辑 SQL 后再执行。
    """
    t_start = time.time()
    logger.info(f"生成 SQL: {request.question[:80]}...")
    try:
        schema = get_schema()
        sql = generate_sql(request.question, schema)

        if sql.upper().strip() == "UNANSWERABLE":
            return GenerateResponse(
                sql="",
                error="抱歉，该问题无法转换为 SQL 查询。",
            )

        if not is_safe_sql(sql):
            return GenerateResponse(
                sql=sql,
                error="⚠ 生成的 SQL 包含危险操作（DROP/DELETE/UPDATE/INSERT/ALTER/TRUNCATE），已被拦截。",
            )

        formatted = format_sql(sql)
        elapsed = time.time() - t_start
        logger.info(f"SQL 生成完成 ({elapsed:.1f}s): {formatted[:100]}")
        return GenerateResponse(sql=formatted, error=None)

    except HTTPException:
        raise
    except Exception as e:
        elapsed = time.time() - t_start
        logger.error(f"SQL 生成失败 ({elapsed:.1f}s): {e}")
        return GenerateResponse(sql="", error=f"生成失败: {str(e)}")


# ==================== SQL 流式生成（SSE） ====================

@app.post("/generate-sql-stream")
async def generate_sql_stream(request: GenerateRequest):
    """
    流式生成 SQL，逐 token 通过 SSE（Server-Sent Events）推送。
    前端可实时看到 SQL 逐字生成，体验类似 ChatGPT。
    """
    from app.llm import generate_sql_stream as gen_stream

    logger.info(f"流式生成 SQL: {request.question[:80]}...")

    async def event_stream():
        try:
            schema = get_schema()
            for token in gen_stream(request.question, schema):
                # SSE 格式: data: <content>\n\n
                yield f"data: {json.dumps({'token': token})}\n\n"
            yield f"data: {json.dumps({'done': True})}\n\n"
        except Exception as e:
            logger.error(f"流式生成失败: {e}")
            yield f"data: {json.dumps({'error': str(e)})}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",  # 禁用 nginx 缓冲
        },
    )


# ==================== 自然语言查询（两步或一步完成） ====================

@app.post("/query", response_model=QueryResponse)
async def query(request: QueryRequest):
    """
    自然语言查询接口：
    - 如果提供了 sql 字段，跳过 LLM 生成，直接执行提供的 SQL
    - 否则：获取 Schema → LLM 生成 SQL → 安全性校验 → 执行 → 翻译
    """
    t_start = time.time()
    sql = request.sql.strip() if request.sql else ""
    formatted_sql = ""
    write_mode = request.write_mode

    if write_mode:
        logger.warning("⚠️ 写模式已启用，允许 UPDATE/DELETE/INSERT 操作")

    try:
        if sql:
            # 用户提供了 SQL，跳过生成，直接执行
            logger.info(f"直接执行 SQL: {sql[:100]}")
            formatted_sql = format_sql(sql)
        else:
            # 自动生成 SQL
            schema = get_schema()
            sql = generate_sql(request.question, schema)

            if sql.upper().strip() == "UNANSWERABLE":
                return QueryResponse(
                    sql="", data=[],
                    error="抱歉，该问题无法转换为 SQL 查询。",
                )

            formatted_sql = format_sql(sql)

        # 安全性检查
        if not is_safe_sql(formatted_sql, write_mode=write_mode):
            raise HTTPException(
                status_code=400,
                detail="生成的 SQL 包含危险操作，已被拒绝。如需执行写操作（UPDATE/DELETE/INSERT），请勾选「启用写操作」选项。",
            )

        # 执行 SQL（失败时自动让 LLM 修正一次）
        try:
            data = execute_sql(formatted_sql)
        except RuntimeError as exec_err:
            error_str = str(exec_err)
            logger.warning(f"首次执行失败，尝试 LLM 自动修正: {error_str[:120]}")

            # 尝试让 LLM 修正 SQL
            try:
                schema = get_schema()
                fixed = fix_sql(request.question or sql, schema, formatted_sql, error_str)
                if fixed.upper().strip() == "UNANSWERABLE":
                    raise RuntimeError(f"{error_str}\n\n💡 提示：LLM 无法自动修正，请点击侧边栏表名查看实际列名后手工修改 SQL。")

                if not is_safe_sql(fixed, write_mode=write_mode):
                    raise RuntimeError(f"{error_str}\n\n💡 修正后的 SQL 仍包含危险操作，已被拦截。")

                fixed_formatted = format_sql(fixed)
                logger.info(f"LLM 修正后的 SQL: {fixed_formatted[:150]}")
                data = execute_sql(fixed_formatted)
                formatted_sql = fixed_formatted  # 使用修正后的 SQL
            except RuntimeError:
                raise  # 重新抛出修正失败的错误
            except Exception as fix_err:
                raise RuntimeError(f"{error_str}\n\n💡 自动修正失败: {fix_err}")

        elapsed = time.time() - t_start
        logger.info(f"查询完成 ({elapsed:.1f}s): {len(data)} 行")

        # 生成中英文翻译
        translation_dict = generate_translation(
            request.question or sql, formatted_sql, data
        )
        translation = Translation(**translation_dict)

        return QueryResponse(
            sql=formatted_sql, data=data,
            translation=translation, error=None,
        )

    except HTTPException:
        raise
    except RuntimeError as e:
        elapsed = time.time() - t_start
        logger.error(f"SQL 执行失败 ({elapsed:.1f}s): {e}")
        return QueryResponse(sql=sql or formatted_sql, data=[], error=str(e))
    except Exception as e:
        elapsed = time.time() - t_start
        logger.error(f"查询失败 ({elapsed:.1f}s): {e}")
        return QueryResponse(
            sql=sql or formatted_sql, data=[],
            error=f"处理请求时发生错误: {str(e)}",
        )


# ==================== 数据导出 ====================

@app.post("/export")
async def export_data(
    sql: str = Form(...),
    format: str = Form("csv"),  # csv / json / xlsx
):
    """
    导出查询结果为文件下载。
    - sql: 要执行的 SQL 查询
    - format: csv / json / xlsx
    返回文件流下载。
    """
    # 安全校验
    if not is_safe_sql(sql):
        raise HTTPException(status_code=400, detail="不允许导出危险操作")

    logger.info(f"导出数据: format={format}, sql={sql[:100]}")
    try:
        data = execute_sql(sql)
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))

    if not data:
        raise HTTPException(status_code=400, detail="查询无数据可导出")

    columns = list(data[0].keys())

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        for row in data:
            writer.writerow([row.get(c, "") for c in columns])
        csv_bytes = output.getvalue().encode("utf-8-sig")
        logger.info(f"导出 CSV: {len(data)} 行")
        return StreamingResponse(
            io.BytesIO(csv_bytes),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=query_result.csv"},
        )

    elif format == "json":
        json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode("utf-8")
        logger.info(f"导出 JSON: {len(data)} 行")
        return StreamingResponse(
            io.BytesIO(json_bytes),
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=query_result.json"},
        )

    elif format == "xlsx":
        # Excel 导出（需要 openpyxl）
        try:
            from openpyxl import Workbook
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="Excel 导出需要 openpyxl 库，请运行 pip install openpyxl",
            )
        wb = Workbook()
        ws = wb.active
        ws.title = "查询结果"

        # 写表头
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 写数据
        for row_idx, row in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        logger.info(f"导出 XLSX: {len(data)} 行")
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": "attachment; filename=query_result.xlsx"
            },
        )
    else:
        raise HTTPException(status_code=400, detail=f"不支持的导出格式: {format}")


# ==================== CSV 导入 ====================

@app.post("/import-csv")
async def import_csv_endpoint(
    file: UploadFile = File(...),
    table_name: str = Form(...),
):
    """导入 CSV 文件并自动创建数据库表（自动检测编码和分隔符）"""
    logger.info(f"导入 CSV: 表名={table_name}")
    try:
        content = await file.read()
        encoding = detect_encoding(content)
        logger.info(f"检测到编码: {encoding}")
        csv_text = content.decode(encoding)
        result = import_csv(table_name, csv_text)
        logger.info(f"导入成功: 表={result['table']}, {result['rows']} 行")
        return {
            "success": True,
            "table": result["table"],
            "rows": result["rows"],
            "columns": result["columns"],
            "message": f"成功导入表 {result['table']}，共 {result['rows']} 行数据",
        }
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="CSV 编码无法识别，请将文件另存为 UTF-8 格式")
    except RuntimeError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# ==================== 表管理 ====================

@app.get("/tables")
async def tables():
    """返回所有表名"""
    from app.database import engine
    from sqlalchemy import inspect
    inspector = inspect(engine)
    return {"tables": inspector.get_table_names()}


@app.delete("/table/{table_name}")
async def delete_table(table_name: str):
    """删除指定表"""
    if not is_safe_table_name(table_name):
        raise HTTPException(status_code=400, detail="表名包含非法字符")
    try:
        info = drop_table(table_name)
        logger.info(f"已删除表: {table_name}, 含 {info['columns']} 列")
        return {"success": True, "message": f"已删除表 '{table_name}'"}
    except RuntimeError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除失败: {str(e)}")
